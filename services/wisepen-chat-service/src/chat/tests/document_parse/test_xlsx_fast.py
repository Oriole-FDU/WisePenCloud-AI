from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image

from chat.application.utils.document_parse.parse_xlsx import fast_parse_xlsx


def test_fast_parse_xlsx_renders_simple_table_without_html_roundtrip(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Name"
    sheet["B1"] = "Logo"
    sheet["A2"] = "Alice"
    sheet["B2"] = '=IMAGE("https://example.com/avatar.png","Avatar")'
    sheet["A3"] = "Docs"
    sheet["B3"] = "WisePen"
    sheet["B3"].hyperlink = "https://example.com/docs"
    file_path = tmp_path / "simple.xlsx"
    workbook.save(file_path)

    markdown = fast_parse_xlsx(file_path)

    assert markdown == (
        "<!-- page 1 -->\n\n"
        "| Name | Logo |\n"
        "| --- | --- |\n"
        "| Alice | ![Avatar](https://example.com/avatar.png) |\n"
        "| Docs | [WisePen](https://example.com/docs) |"
    )
    assert "<table>" not in markdown


def test_fast_parse_xlsx_does_not_duplicate_islands_inside_table_bbox(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    for coordinate in (
        "A1",
        "B1",
        "C1",
        "D1",
        "E1",
        "A2",
        "A3",
        "A4",
        "A5",
        "E2",
        "E3",
        "E4",
        "E5",
        "B5",
        "C5",
        "D5",
    ):
        sheet[coordinate] = "1"
    sheet["C3"] = "island"
    file_path = tmp_path / "island.xlsx"
    workbook.save(file_path)

    markdown = fast_parse_xlsx(file_path)

    assert markdown.count("island") == 1


def test_fast_parse_xlsx_preserves_merged_cells_as_html_and_sheet_pages(tmp_path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "Merged"
    first.merge_cells("A1:B1")
    first["A1"] = "Header"
    first["A2"] = "A"
    first["B2"] = "B"
    second = workbook.create_sheet("Single")
    second["A1"] = "Only text"
    hidden = workbook.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "hidden text"
    file_path = tmp_path / "merged.xlsx"
    workbook.save(file_path)

    markdown = fast_parse_xlsx(file_path)

    assert "<!-- page 1 -->\n\n# Merged\n\n<table>" in markdown
    assert '<th colspan="2">Header</th>' in markdown
    assert "<!-- page 2 -->\n\n# Single\n\nOnly text" in markdown
    assert "hidden text" not in markdown


def test_fast_parse_xlsx_extracts_anchored_sheet_images(tmp_path: Path) -> None:
    source_image = tmp_path / "source.png"
    Image.new("RGB", (2, 2), color="red").save(source_image)
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Table"
    sheet.add_image(OpenpyxlImage(source_image), "C3")
    file_path = tmp_path / "image.xlsx"
    workbook.save(file_path)

    markdown = fast_parse_xlsx(file_path, image_path=tmp_path / "images")

    assert "Table" in markdown
    assert "![image1.png](images/image1.png)" in markdown
    assert (tmp_path / "images" / "image1.png").is_file()
