from __future__ import annotations

import html
import re
from collections import deque
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

_IMAGE_FORMULA_RE = re.compile(r"^=\s*(?:_xlfn\.)?IMAGE\s*\((?P<args>.*)\)\s*$", re.I)
_FORMULA_STRING_RE = re.compile(r'"((?:[^"]|"")*)"')


@dataclass(frozen=True, slots=True)
class CellRender:
    markdown: str
    html: str


@dataclass(frozen=True, slots=True)
class TableCell:
    row: int
    col: int
    render: CellRender
    row_span: int = 1
    col_span: int = 1


@dataclass(frozen=True, slots=True)
class SheetBlock:
    row: int
    col: int
    order: int
    markdown: str


class XlsxConverter:
    """直接从 workbook 结构渲染 XLSX，避免 MinerU HTML round-trip。"""

    def convert(
        self,
        file_path: str | Path,
        *,
        image_path: str | Path | None = None,
    ) -> str:
        file_path = Path(file_path)
        if not file_path.is_file():
            raise FileNotFoundError(file_path)

        image_path = Path(image_path) if image_path is not None else None
        value_workbook = load_workbook(file_path, data_only=True, rich_text=True)
        formula_workbook = load_workbook(file_path, data_only=False, rich_text=True)
        try:
            pages = self._render_workbook(
                value_workbook.worksheets,
                formula_workbook.worksheets,
                image_path,
            )
        finally:
            value_workbook.close()
            formula_workbook.close()

        return "\n\n".join(
            f"<!-- page {index} -->\n\n{page}" if page else f"<!-- page {index} -->"
            for index, page in enumerate(pages, start=1)
        )

    def _render_workbook(
        self,
        value_sheets: list[Worksheet],
        formula_sheets: list[Worksheet],
        image_path: Path | None,
    ) -> list[str]:
        sheet_pages: list[tuple[str, list[str]]] = []
        for value_sheet, formula_sheet in zip(value_sheets, formula_sheets, strict=True):
            if value_sheet.sheet_state != Worksheet.SHEETSTATE_VISIBLE:
                continue
            blocks = self._render_sheet(value_sheet, formula_sheet, image_path)
            if blocks:
                sheet_pages.append((value_sheet.title, blocks))

        # 只有多个非空 sheet 时才输出标题，
        # 避免单 sheet 文件多一个没有检索价值的 heading。
        add_sheet_titles = len(sheet_pages) > 1
        pages: list[str] = []
        for title, blocks in sheet_pages:
            page_blocks = [f"# {title}", *blocks] if add_sheet_titles else blocks
            pages.append("\n\n".join(page_blocks))
        return pages

    def _render_sheet(
        self,
        value_sheet: Worksheet,
        formula_sheet: Worksheet,
        image_path: Path | None,
    ) -> list[str]:
        lookup = _MergedCellLookup(value_sheet)
        components = _find_table_components(value_sheet, formula_sheet, lookup)
        blocks = [
            SheetBlock(row, col, index, markdown)
            for index, (row, col, markdown) in enumerate(
                _render_table_component(value_sheet, formula_sheet, lookup, component)
                for component in components
            )
            if markdown
        ]
        blocks.extend(_render_floating_images(value_sheet, image_path, len(blocks)))
        return [
            block.markdown
            for block in sorted(blocks, key=lambda item: (item.row, item.col, item.order))
        ]


class _MergedCellLookup:
    """按坐标缓存合并单元格，蒸馏自 MinerU 的 merged-cell lookup。"""

    def __init__(self, sheet: Worksheet) -> None:
        self._hidden: set[tuple[int, int]] = set()
        self._spans: dict[tuple[int, int], tuple[int, int]] = {}
        self._merged: set[tuple[int, int]] = set()

        for merged in sheet.merged_cells.ranges:
            min_row = merged.min_row - 1
            max_row = merged.max_row - 1
            min_col = merged.min_col - 1
            max_col = merged.max_col - 1
            self._spans[(min_row, min_col)] = (
                max_row - min_row + 1,
                max_col - min_col + 1,
            )
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    self._merged.add((row, col))
                    if (row, col) != (min_row, min_col):
                        self._hidden.add((row, col))

    def contains(self, row: int, col: int) -> bool:
        return (row, col) in self._merged

    def is_hidden(self, row: int, col: int) -> bool:
        return (row, col) in self._hidden

    def span(self, row: int, col: int) -> tuple[int, int]:
        return self._spans.get((row, col), (1, 1))


def _find_table_components(
    value_sheet: Worksheet,
    formula_sheet: Worksheet,
    lookup: _MergedCellLookup,
) -> list[set[tuple[int, int]]]:
    # 蒸馏 MinerU 的表格发现思路：用内容单元格和合并区域做图节点，
    # 再按四邻域连通分量形成表格，避免先渲染成 HTML 后再反解析。
    content_cells = {
        (row - 1, col - 1)
        for row in range(1, max(value_sheet.max_row, formula_sheet.max_row) + 1)
        for col in range(1, max(value_sheet.max_column, formula_sheet.max_column) + 1)
        if _cell_has_content(value_sheet, formula_sheet, lookup, row - 1, col - 1)
    }

    components: list[set[tuple[int, int]]] = []
    visited: set[tuple[int, int]] = set()
    for start in sorted(content_cells):
        if start in visited:
            continue
        component: set[tuple[int, int]] = set()
        queue = deque([start])
        visited.add(start)
        while queue:
            row, col = queue.popleft()
            component.add((row, col))
            for neighbor in (
                (row + 1, col),
                (row - 1, col),
                (row, col + 1),
                (row, col - 1),
            ):
                if neighbor in content_cells and neighbor not in visited:
                    visited.add(neighbor)
                    queue.append(neighbor)
        components.append(component)
    return _merge_overlapping_components(components)


def _merge_overlapping_components(
    components: list[set[tuple[int, int]]],
) -> list[set[tuple[int, int]]]:
    merged: list[set[tuple[int, int]]] = []
    for component in components:
        combined = set(component)
        index = 0
        while index < len(merged):
            if _bbox_overlaps(_component_bbox(combined), _component_bbox(merged[index])):
                combined |= merged.pop(index)
                index = 0
                continue
            index += 1
        merged.append(combined)
    return merged


def _component_bbox(component: set[tuple[int, int]]) -> tuple[int, int, int, int]:
    rows = [row for row, _ in component]
    cols = [col for _, col in component]
    return min(rows), min(cols), max(rows), max(cols)


def _bbox_overlaps(
    first: tuple[int, int, int, int],
    second: tuple[int, int, int, int],
) -> bool:
    first_min_row, first_min_col, first_max_row, first_max_col = first
    second_min_row, second_min_col, second_max_row, second_max_col = second
    return (
        first_min_row <= second_max_row
        and second_min_row <= first_max_row
        and first_min_col <= second_max_col
        and second_min_col <= first_max_col
    )


def _cell_has_content(
    value_sheet: Worksheet,
    formula_sheet: Worksheet,
    lookup: _MergedCellLookup,
    row: int,
    col: int,
) -> bool:
    value = value_sheet.cell(row=row + 1, column=col + 1).value
    formula = formula_sheet.cell(row=row + 1, column=col + 1).value
    return (
        value not in (None, "")
        or formula not in (None, "")
        or lookup.contains(row, col)
    )


def _render_table_component(
    value_sheet: Worksheet,
    formula_sheet: Worksheet,
    lookup: _MergedCellLookup,
    component: set[tuple[int, int]],
) -> tuple[int, int, str]:
    min_row = min(row for row, _ in component)
    max_row = max(row for row, _ in component)
    min_col = min(col for _, col in component)
    max_col = max(col for _, col in component)
    table_cells: list[TableCell] = []

    # bbox 内部的空单元格需要保留，否则 Markdown 表格会丢列；
    # 只有合并区域的非左上角隐藏格才跳过。
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if lookup.is_hidden(row, col):
                continue
            row_span, col_span = lookup.span(row, col)
            table_cells.append(
                TableCell(
                    row=row - min_row,
                    col=col - min_col,
                    render=_render_cell(
                        value_sheet.cell(row=row + 1, column=col + 1),
                        formula_sheet.cell(row=row + 1, column=col + 1),
                    ),
                    row_span=row_span,
                    col_span=col_span,
                )
            )

    if (
        len(table_cells) == 1
        and table_cells[0].row_span == 1
        and table_cells[0].col_span == 1
    ):
        return min_row, min_col, table_cells[0].render.markdown

    has_merge = any(cell.row_span > 1 or cell.col_span > 1 for cell in table_cells)
    if has_merge:
        markdown = _render_html_table(
            table_cells,
            max_row - min_row + 1,
            max_col - min_col + 1,
        )
    else:
        markdown = _render_pipe_table(
            table_cells,
            max_row - min_row + 1,
            max_col - min_col + 1,
        )
    return min_row, min_col, markdown


def _render_cell(value_cell, formula_cell) -> CellRender:
    # IMAGE 公式需要从公式工作簿读取；data_only 工作簿只负责普通单元格显示值。
    image = _image_formula(formula_cell.value)
    if image is not None:
        url, alt = image
        return CellRender(
            markdown=f"![{_escape_link_label(alt)}]({url})",
            html=(
                f'<img src="{html.escape(url, quote=True)}" '
                f'alt="{html.escape(alt, quote=True)}" />'
            ),
        )

    text = "" if value_cell.value is None else str(value_cell.value)
    target = _hyperlink_target(value_cell) or _hyperlink_target(formula_cell)
    if target and text:
        escaped_text = html.escape(text).replace("\n", "<br>")
        return CellRender(
            markdown=f"[{_escape_link_label(text)}]({target})",
            html=f'<a href="{html.escape(target, quote=True)}">{escaped_text}</a>',
        )

    return CellRender(
        markdown=_escape_markdown_cell(text),
        html=html.escape(text).replace("\n", "<br>"),
    )


def _image_formula(value: object) -> tuple[str, str] | None:
    if not isinstance(value, str):
        return None
    match = _IMAGE_FORMULA_RE.match(value.strip())
    if match is None:
        return None
    args = [
        item.replace('""', '"')
        for item in _FORMULA_STRING_RE.findall(match.group("args"))
    ]
    if not args:
        return None
    url = _safe_url(args[0])
    if not url:
        return None
    return url, args[1] if len(args) > 1 and args[1] else "image"


def _safe_url(value: str) -> str:
    url = value.strip()
    scheme = urlparse(url).scheme.lower()
    return url if scheme in {"http", "https"} else ""


def _hyperlink_target(cell) -> str:
    hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink is None:
        return ""
    target = getattr(hyperlink, "target", None) or getattr(hyperlink, "location", None)
    if target is None:
        return ""
    target = str(target).strip()
    if not target:
        return ""
    if getattr(hyperlink, "location", None) and not getattr(hyperlink, "target", None):
        return f"#{target}"
    return target


def _render_pipe_table(cells: list[TableCell], row_count: int, col_count: int) -> str:
    matrix = [["" for _ in range(col_count)] for _ in range(row_count)]
    for cell in cells:
        matrix[cell.row][cell.col] = cell.render.markdown
    header = matrix[0] if matrix else []
    rows = matrix[1:] if len(matrix) > 1 else []
    return "\n".join(
        [
            "| " + " | ".join(header) + " |",
            "| " + " | ".join("---" for _ in header) + " |",
            *("| " + " | ".join(row) + " |" for row in rows),
        ]
    )


def _render_html_table(cells: list[TableCell], row_count: int, col_count: int) -> str:
    cell_map = {(cell.row, cell.col): cell for cell in cells}
    covered: set[tuple[int, int]] = set()
    lines = ["<table>"]
    for row in range(row_count):
        lines.append("<tr>")
        for col in range(col_count):
            if (row, col) in covered:
                continue
            cell = cell_map.get((row, col))
            if cell is None:
                lines.append("<td></td>")
                continue
            for row_offset in range(cell.row_span):
                for col_offset in range(cell.col_span):
                    covered.add((row + row_offset, col + col_offset))
            tag = "th" if row == 0 else "td"
            attrs = []
            if cell.row_span > 1:
                attrs.append(f'rowspan="{cell.row_span}"')
            if cell.col_span > 1:
                attrs.append(f'colspan="{cell.col_span}"')
            attr_text = " " + " ".join(attrs) if attrs else ""
            lines.append(f"<{tag}{attr_text}>{cell.render.html}</{tag}>")
        lines.append("</tr>")
    lines.append("</table>")
    return "\n".join(lines)


def _render_floating_images(
    sheet: Worksheet,
    image_path: Path | None,
    order_start: int,
) -> list[SheetBlock]:
    if image_path is None:
        return []
    blocks: list[SheetBlock] = []
    for index, image in enumerate(getattr(sheet, "_images", []), start=1):
        # 第一版只处理 openpyxl 已暴露 anchor 的浮动图，不解析 shape/chart/OCR。
        row, col = _image_anchor(image)
        suffix = (getattr(image, "format", None) or "png").lower()
        image_name = f"image{index}.{suffix}"
        image_path.mkdir(parents=True, exist_ok=True)
        (image_path / image_name).write_bytes(image._data())
        blocks.append(
            SheetBlock(
                row=row,
                col=col,
                order=order_start + index,
                markdown=f"![{image_name}]({image_path.name}/{image_name})",
            )
        )
    return blocks


def _image_anchor(image) -> tuple[int, int]:
    marker = getattr(getattr(image, "anchor", None), "_from", None)
    if marker is None:
        return 10**9, 10**9
    return int(marker.row), int(marker.col)


def _escape_markdown_cell(value: str) -> str:
    return (
        value.replace("\\", "\\\\")
        .replace("|", "\\|")
        .replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("\n", "<br>")
    )


def _escape_link_label(value: str) -> str:
    return value.replace("[", "\\[").replace("]", "\\]").replace("\n", " ")
