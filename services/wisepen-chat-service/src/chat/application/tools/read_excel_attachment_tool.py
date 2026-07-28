from typing import Dict, Any

import httpx

from chat.core.config.app_settings import settings
from chat.application.tools.core import (
    ToolDefinition,
    ToolExecutionError,
    ToolLLMSpec,
    ToolParametersSchema,
    ToolPolicy,
    ToolRiskLevel,
)
from chat.domain.repositories import SessionRepository
from chat.service_client.file_storage_service_client import FileStorageClient


class ReadExcelAttachmentTool:
    """Extract text content from an Excel (.xlsx) attachment."""

    def __init__(self, file_storage_client: FileStorageClient, session_repo: SessionRepository) -> None:
        self._file_storage = file_storage_client
        self._session_repo = session_repo
        parameters_schema: Dict[str, Any] = {
            "type": "object",
            "properties": {
                "attachment_id": {
                    "type": "string",
                    "description": "The attachment ID to read.",
                },
            },
            "required": ["attachment_id"],
        }
        self._definition = ToolDefinition(
            llm_spec=ToolLLMSpec(
                name="read_excel_attachment",
                description=(
                    "Extract and read the text content of an Excel (.xlsx or .xlsm) attachment. "
                    "Returns a TSV-like text representation of each sheet. Use this for .xlsx files attached to the conversation."
                ),
                parameters_schema=ToolParametersSchema(parameters_schema),
            ),
            policy=ToolPolicy(
                expose_by_default=True,
                persist_output=False,
                risk_level=ToolRiskLevel.LOW,
                required_context_keys=("user_id", "session_id"),
                timeout_seconds=30.0,
                max_output_chars=settings.TOOL_RESULT_MAX_CHARS,
            ),
        )

    @property
    def definition(self) -> ToolDefinition:
        return self._definition

    async def execute(
        self,
        context: dict[str, Any],
        config: dict[str, Any] | None = None,
        **kwargs: Any,
    ) -> str:
        user_id = context.get("user_id")
        session_id = context.get("session_id")
        attachment_id = kwargs.get("attachment_id", "").strip()
        if not attachment_id:
            raise ToolExecutionError(reason="missing_attachment_id", detail_reason="Missing required argument: attachment_id.")

        temp_refs, _res_refs = await self._session_repo.get_session_attachments(session_id, user_id)
        ref = next((r for r in (temp_refs or []) if r.attachment_id == attachment_id), None)
        if ref is None:
            raise ToolExecutionError(reason="attachment_not_found", detail_reason=f"Attachment {attachment_id} not found in session.")

        download_url = await self._file_storage.get_download_url(ref.object_key)
        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.get(download_url)
                resp.raise_for_status()
                xlsx_bytes = resp.content
        except httpx.HTTPError as e:
            raise ToolExecutionError(reason="download_failed", detail_reason=f"Failed to download attachment: {e}") from e

        try:
            import openpyxl
            import io
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
            sheet_texts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append("\t".join(str(cell) if cell is not None else "" for cell in row))
                sheet_texts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
            wb.close()
            text = "\n\n".join(sheet_texts)
        except ImportError:
            raise ToolExecutionError(reason="xlsx_parser_unavailable", detail_reason="Excel parser (openpyxl) is not installed.")
        except Exception as e:
            raise ToolExecutionError(reason="xlsx_parse_failed", detail_reason=f"Failed to parse Excel file: {e}") from e

        if len(text) > settings.TOOL_RESULT_MAX_CHARS:
            text = text[:settings.TOOL_RESULT_MAX_CHARS] + "\n...[truncated]"
        return text
